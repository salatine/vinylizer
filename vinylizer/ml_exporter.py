from typing import List

from openpyxl.cell import MergedCell
from openpyxl.worksheet import worksheet
from vinylizer.Product import Product
import openpyxl

HEADER_ROW = 3
FIELDS_BY_CELL_VALUE = {
    'Forma de anunciar': 'listing-type',
    'Título': 'title',
    'Código universal': 'code',
    'Fotos': 'picture_urls',
    'Estoque': 'stock',
    'Canal de venda': 'channel',
    'Condição': 'condition',
    'Descrição': 'description',
    'Tipo de anúncio': 'type',
    'Forma de envio': 'shipping-mode',
    'Retirar pessoalmente': 'pickup',
    'Tipo de garantia': 'warranty',
    'Nome do artista': 'artist',
    'Nome do álbum': 'album',
    'Companhia': 'label',
    'Tipo de álbum': 'album-type',
    'Formato': 'format',
    'Incluí faixas': 'tracks',
    'Ano de': 'year',
    'embalagem': 'packaging',
    'canções': 'songs',
    'Origem': 'nationality',
    'Gênero': 'genre',
    'Acessórios': 'accessories',
    'peças': 'pieces',
    'Preço': 'price-ml',
    'Custo de envio': 'shipping-ml',
    'Largura (cm)': 'width',
    'Altura (cm)': 'height',
    'Profundidade (cm)': 'depth',
    'Peso físico (kg)': 'weight',
}

MEASURES_BY_FORMAT = {
    'Lp Vinil': {
        'width': 40,
        'height': 5,
        'depth': 40,
        'weight': 1,
    },
    # TODO: measurements for these others formats
    'Compacto Vinil': {
        'width': 0,
        'height': 0,
        'depth': 0,
        'weight': 0,
    },
    'CD': {
        'width': 0,
        'height': 0,
        'depth': 0,
        'weight': 0,
    },
    'DVD': {
        'width': 0,
        'height': 0,
        'depth': 0,
        'weight': 0,
    },
    'Fita K0 Cassete': {
        'width': 0,
        'height': 0,
        'depth': 0,
        'weight': 0,
    },
    'LD LaserDisc': {
        'width': 0,
        'height': 0,
        'depth': 0,
        'weight': 0,
    },
}

def export_to_ml_spreadsheet(spreadsheet_path: str, products: List[Product]) -> None:
    wb = openpyxl.load_workbook(spreadsheet_path)
    ws = wb['Música']

    start_row = get_start_row(ws)
    columns_to_write = get_columns_to_write(ws)

    for i, product in enumerate(products):
        print(f'{i+1}/{len(products)}: {product.title}')
        fields = {
            'listing-type': 'Lista geral',
            'title': product.title,
            'code': 'O produto não tem código cadastrado',
            'picture_urls': ', '.join(product.picture_urls),
            'stock': product.stock,
            'channel': 'Mercado Livre',
            'price-ml': product.price,
            'condition': 'Novo' if product.is_new else 'Usado',
            'description': product.description,
            'type': 'Clássico',
            'shipping-mode': 'Mercado Envios | Mercado Envios Flex',
            'shipping-ml': 'Por conta do comprador',
            'pickup': 'Não aceito',
            'warranty': 'Sem garantia',
            'artist': product.artist if product.artist else product.album,
            'album': product.album if product.album.strip() != '' and product.album is not None else product.artist,
            'label': product.label or 'N/A',
            'album-type': 'Vinil' if 'Vinil' in product.format else product.format,
            'format': 'Físico',
            'tracks': 'Não',
            'year': product.release_year or 'N/A',
            'packaging': 'N/A',
            'songs': product.song_quantity,
            'nationality': product.nationality_text.title(),
            'genre': product.genres[0].title(),
            'accessories': 'N/A',
            'pieces': 1,
            'album-duration': product.album_duration,
            'album-duration-time-unit': 'm',
            'width': MEASURES_BY_FORMAT[product.format]["width"],
            'height': MEASURES_BY_FORMAT[product.format]["height"],
            'depth': MEASURES_BY_FORMAT[product.format]["depth"],
            'weight': MEASURES_BY_FORMAT[product.format]["weight"],
        }

        for field, column in columns_to_write.items():
            ws[f'{column}{start_row + i}'] = fields[field]

    wb.save(spreadsheet_path)

def get_start_row(ws: worksheet.Worksheet) -> int:
    for i in range(HEADER_ROW, ws.max_row):
        cell = ws.cell(row=i, column=4)
        if 'Novo' in str(cell.value):
            return cell.row
    raise Exception('Não foi possível encontrar a linha de início')

def get_columns_to_write(ws: worksheet.Worksheet) -> dict:
    columns_to_write = {}
    for column in ws.iter_cols(min_row=HEADER_ROW, max_row=HEADER_ROW + 1):
        for cell in column:
            if cell.value is None:
                continue

            value = str(cell.value)
            for cell_value, field in FIELDS_BY_CELL_VALUE.items():
                if cell_value in value:
                    columns_to_write[field] = cell.column_letter

            if 'Duração total do álbum' in value:
                if 'Duração total do álbum' == value:
                    columns_to_write['album-duration'] = cell.column_letter
                else:
                    columns_to_write['album-duration-time-unit'] = cell.column_letter

    return columns_to_write

def parser_merged_cell(sheet: worksheet.Worksheet, row, col):
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # return the left top cell
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell
