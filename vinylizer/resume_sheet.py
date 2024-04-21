import openpyxl
from openpyxl import utils, styles
from vinylizer.Product import Product
from typing import List

def columns_best_fit(ws):
    for column_cells in ws.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (utils.get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            ws.column_dimensions[new_column_letter].width = new_column_length*1.23

def create_resume_sheet(products: List[Product], resume_book_path):
    resume_book = openpyxl.Workbook()
    resume_sheet = resume_book[resume_book.sheetnames[0]]

    resume_sheet['A1'] = 'Título'
    resume_sheet['B1'] = 'Preço'
    resume_sheet['C1'] = 'Plataformas'
    
    black = styles.Side(style='thin', color='000000')
    sides_black_border = styles.Border(left=black, right=black)
    top_black_border = styles.Border(left=black, right=black, top=black)
    bottom_black_border = styles.Border(left=black, right=black, bottom=black)
    center = styles.Alignment(horizontal='center')

    for i, product in enumerate(products):
        resume_sheet[f'A{i + 2}'] = product.title
        resume_sheet[f'B{i + 2}'] = product.price # type: ignore
        resume_sheet[f'B{i + 2}'].number_format = '0.00'
        resume_sheet[f'C{i + 2}'] = 'ML, Shopify'

        for letter in ['A', 'B', 'C']:
            resume_sheet[f'{letter}{i + 2}'].border = sides_black_border
            resume_sheet[f'{letter}{i + 2}'].alignment = center
            if i == 0:
                resume_sheet[f'{letter}{i + 2}'].border = top_black_border
            if i == len(products) - 1:
                resume_sheet[f'{letter}{i + 2}'].border = bottom_black_border
                resume_sheet[f'{letter}1'].border = top_black_border
                resume_sheet[f'{letter}1'].alignment = center
                resume_sheet[f'{letter}{len(products) + 2}'].border = bottom_black_border
                resume_sheet[f'{letter}{len(products) + 2}'].alignment = center

    resume_sheet[f'A{len(products) + 2}'] = len(products) # type: ignore
    resume_sheet[f'B{len(products) + 2}'] = f'=SUM(B2:B{len(products) + 1})'
    resume_sheet[f'C{len(products) + 2}'] = 'ML, Shopify'

    columns_best_fit(resume_sheet)

    resume_book.save(resume_book_path)

